import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Mott MacDonald SOC ticket Processor", layout="wide")
st.title("📊 Mott MacDonald SOC ticket Processor")

# === File Upload ===
incident_file = st.file_uploader("Upload ServiceNow SOC incident Report.xlsx", type=["xlsx"])
autodesk_file = st.file_uploader("Upload tenable report.csv", type=["csv"])

if incident_file and autodesk_file:
    try:
        # === Load Incident File ===
        df_incident = pd.read_excel(incident_file)

        # Extract solution(s) from Description
        def extract_solutions(text):
            if pd.isna(text):
                return []
            # Format 1: Solution: ... later.
            match_single = re.search(r'Solution:\s*(.*?later\.)', text, flags=re.IGNORECASE)
            if match_single:
                return [match_single.group(1).strip()]
            # Format 2: Solution(s): [_..._,_..._]
            match_multi = re.search(r'Solution\(s\):\s*\[(.*?)\]', text, flags=re.IGNORECASE)
            if match_multi:
                raw = match_multi.group(1)
                return [s.strip('_') for s in raw.split('_,_') if s.strip('_')]
            return []

        solution_columns = df_incident['Description'].apply(extract_solutions).apply(pd.Series)
        solution_columns.columns = [f"Solution {i+1}" for i in range(solution_columns.shape[1])]
        df_incident = pd.concat([df_incident, solution_columns], axis=1)

        # Extract hosts from Description
        def extract_hosts(text):
            if pd.isna(text):
                return []
            match = re.search(r'\[(.*?)\]', text)
            if match:
                return [host.strip('_') for host in match.group(1).split(',') if host.strip('_')]
            return []

        host_columns = df_incident['Description'].apply(extract_hosts).apply(pd.Series)
        host_columns.columns = [f"Host {i+1}" for i in range(host_columns.shape[1])]
        df_incident = pd.concat([df_incident, host_columns], axis=1)

        # === Load Autodesk File ===
        df_autodesk = pd.read_csv(autodesk_file)
        drop_cols = [
            "asset.display_fqdn", "asset.display_ipv4_address", "asset.id", "asset.tags",
            "asset_inventory", "definition.cvss3.base_score", "definition.description",
            "definition.family", "definition.id", "definition.patch_published",
            "definition.plugin_published", "definition.plugin_updated", "definition.vpr.score",
            "definition.vulnerability_published", "id", "port", "protocol"
        ]
        df_autodesk.drop(columns=drop_cols, inplace=True, errors='ignore')
        df_autodesk["application"] = df_autodesk["output"].str.extract(
            r'Autodesk\\(.*?)(?:\\|\s)+Installed version', expand=False)

        if "state" in df_autodesk.columns:
            df_autodesk = df_autodesk[df_autodesk["state"].str.lower() != "fixed"]
            df_autodesk.drop(columns=["state"], inplace=True)

        df_autodesk.dropna(subset=["asset.name", "definition.solution"], inplace=True)
        df_autodesk["asset.name"] = df_autodesk["asset.name"].str.lower().str.strip()
        df_autodesk["definition.solution"] = df_autodesk["definition.solution"].str.strip()

        # === Apply Color Coding ===
        output = BytesIO()
        df_incident.to_excel(output, index=False)
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        host_cols = [col for col in df_incident.columns if col.startswith("Host ")]
        solution_cols = [col for col in df_incident.columns if col.startswith("Solution ")]
        green_only_rows = []
        host_occurrences = {}

        for idx, row in df_incident.iterrows():
            valid_hosts = []
            green_hosts = []

            for col in host_cols:
                host = str(row.get(col, "")).strip().lower()
                if not host:
                    continue
                valid_hosts.append(host)
                host_occurrences.setdefault(host, []).append(idx)
                col_index = df_incident.columns.get_loc(col) + 1
                cell = ws.cell(row=idx + 2, column=col_index)

                matched = False
                for sol_col in solution_cols:
                    sol = str(row.get(sol_col, "")).strip()
                    if not sol:
                        continue
                    match = df_autodesk[
                        (df_autodesk["asset.name"] == host) &
                        (df_autodesk["definition.solution"] == sol)
                    ]
                    if not match.empty:
                        cell.fill = fill_red
                        green_hosts.append(host)
                        matched = True
                        break
                if not matched:
                    cell.fill = fill_green

            if green_hosts:
                green_only_rows.append({
                    "Number": row.get("Number", ""),
                    "Matched Hosts": ", ".join(green_hosts),
                    "Solutions": ", ".join([str(row.get(col, "")).strip() for col in solution_cols if row.get(col)])
                })

            # Write other cells without coloring
            for col_num, col_name in enumerate(df_incident.columns, start=1):
                if col_name not in host_cols:
                    ws.cell(row=idx + 2, column=col_num, value=row.get(col_name, ""))

        # === Build Clean Duplicate Host Report with Matching Solutions ===
        duplicate_host_summary = []
        for host, indices in host_occurrences.items():
            if len(indices) > 1:
                filtered = df_incident.loc[indices]
                filtered = filtered.dropna(subset=["Number"])
                all_solutions = filtered[solution_cols].fillna("").agg(','.join, axis=1).str.strip()
                solution_set = set(all_solutions)
                if len(solution_set) == 1:
                    incident_numbers = sorted(set(filtered["Number"].astype(str)))
                    duplicate_host_summary.append({
                        "Duplicate Host": host,
                        "Incident Numbers": ", ".join(incident_numbers),
                        "Shared Solution Matching": list(solution_set)[0]
                    })

        df_duplicates = pd.DataFrame(duplicate_host_summary)

        # === Save Final Excel ===
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # === Streamlit Outputs ===
        st.success("✅ Processing complete!")
        st.download_button("📥 Download Colored Incident File", data=final_output, file_name="incident_with_hosts_colored.xlsx")
        st.download_button("📥 Download Modified Autodesk CSV", data=df_autodesk.to_csv(index=False).encode("utf-8"), file_name="Modified_Autodesk_Report.csv")

        with st.expander("✅ Incidents that require troubleshooting"):
            if green_only_rows:
                st.dataframe(pd.DataFrame(green_only_rows), use_container_width=True)
            else:
                st.write("No incidents that require troubleshooting.")

        with st.expander("🔁 Duplicate Hosts Across Rows"):
            if not df_duplicates.empty:
                st.dataframe(df_duplicates, use_container_width=True)
            else:
                st.write("No duplicate hosts found with matching solutions.")

    except Exception as e:
        st.error(f"❌ Error during processing: {e}")
else:
    st.info("📂 Please upload both files to begin.")
